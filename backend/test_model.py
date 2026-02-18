#!/usr/bin/env python3
"""
Model evaluation script for Teacher's Pet math tutor AI.

Reads test cases from the evaluation xlsx, queries either Gemini or GitHub Models API using the
same system prompt as production, and writes responses into the "Actual Output"
column of each sheet. Saves results as a new xlsx for manual scoring.

Usage:
    # Gemini models
    python test_model.py --input ../math_tutor_ai_eval_testcases.xlsx --model gemini-2.5-flash-lite --provider gemini
    
    # GitHub Models
    python test_model.py --input ../math_tutor_ai_eval_testcases.xlsx --model gpt-4o --provider github
    python test_model.py --input ../math_tutor_ai_eval_testcases.xlsx --model meta-llama-3.1-405b-instruct --provider github
"""

import argparse
import os
import re
import time
from datetime import date

from dotenv import load_dotenv
from google import genai
import openpyxl

load_dotenv()

# Must match production system prompt exactly for fair evaluation
SYSTEM_PROMPT = """You are Teacher's Pet, a friendly and encouraging math tutor for middle school students (grades 6-8).

TEACHING STYLE:
- Always work through problems step-by-step, numbering each step clearly
- Use simple, age-appropriate language — define any math terms you use
- Guide students toward answers with hints and questions rather than just stating the answer
- When a student makes a mistake, gently explain what went wrong and why
- Be warm, patient, and encouraging — acknowledge effort and celebrate progress
- Offer a second explanation method or real-world example when it helps understanding
- End every response by clearly stating the final answer

SAFETY & BOUNDARIES:
- Only help with math topics from the middle school curriculum (Common Core grades 6-8)
- Never ask for or encourage sharing of personal information (name, age, school, location, etc.)
- If a student expresses distress, respond with empathy and encourage them to talk to a trusted adult
- Politely decline off-topic, inappropriate, or unethical requests (e.g. cheating)"""

# Non-test sheets to skip
SKIP_SHEETS = {"Instructions & Rubric", "Summary Dashboard"}

PROMPT_HEADER = "Test Input / Prompt"
OUTPUT_HEADER = "Actual Output"


def find_col(headers, name):
    """Return 1-based column index for a header name, or None if not found."""
    for i, h in enumerate(headers, 1):
        if h and str(h).strip() == name:
            return i
    return None


def query_gemini(client, model, question, retries=3):
    """Send a question to Gemini, auto-retrying on rate limit or overload errors."""
    for attempt in range(retries):
        try:
            response = client.models.generate_content(
                model=model,
                contents=f"{SYSTEM_PROMPT}\n\nStudent: {question}"
            )
            return response.text
        except Exception as e:
            msg = str(e)
            # Parse suggested retry delay from the error message if present
            match = re.search(r'retry[^\d]*(\d+(?:\.\d+)?)\s*s', msg, re.IGNORECASE)
            wait = float(match.group(1)) + 1 if match else 5
            if attempt < retries - 1 and ("429" in msg or "503" in msg):
                print(f"    -> rate limited, waiting {wait:.0f}s then retrying...")
                time.sleep(wait)
            else:
                raise


def query_github(token, model, question, retries=3):
    """Send a question to GitHub Models API."""
    import requests
    
    endpoint = "https://models.inference.ai.azure.com/chat/completions"
    
    for attempt in range(retries):
        try:
            response = requests.post(
                endpoint,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {token}"
                },
                json={
                    "model": model,
                    "messages": [
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user", "content": question}
                    ],
                    "temperature": 0.7,
                    "max_tokens": 2000
                },
                timeout=90
            )
            
            if response.status_code == 200:
                data = response.json()
                return data["choices"][0]["message"]["content"]
            elif response.status_code == 429 and attempt < retries - 1:
                wait = 5
                print(f"    -> rate limited, waiting {wait}s then retrying...")
                time.sleep(wait)
            else:
                response.raise_for_status()
                
        except Exception as e:
            # If this is an HTTP 400 (Bad Request) from the provider, do not retry —
            # these are blocked by safety filters and will not succeed on retry.
            try:
                # Many requests exceptions expose a response with a status_code
                if hasattr(e, 'response') and e.response is not None and getattr(e.response, 'status_code', None) == 400:
                    raise
            except Exception:
                # If we intentionally re-raised, let it bubble up to the caller
                raise

            if attempt < retries - 1:
                print(f"    -> error: {e}, retrying...")
                time.sleep(3)
            else:
                raise


def query_model(client_or_token, model, question, provider="gemini", retries=3):
    """Send a question to the appropriate model provider."""
    if provider == "gemini":
        return query_gemini(client_or_token, model, question, retries)
    elif provider == "github":
        return query_github(client_or_token, model, question, retries)
    else:
        raise ValueError(f"Unknown provider: {provider}")


def count_questions(wb):
    """Count total non-empty prompts across all test sheets that need processing."""
    count = 0
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        prompt_col = find_col(headers, PROMPT_HEADER)
        output_col = find_col(headers, OUTPUT_HEADER)
        if prompt_col is None or output_col is None:
            continue
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=prompt_col).value
            if val and str(val).strip():
                # Only count if output is empty or contains a non-HTTP-400 error
                existing_output = ws.cell(row=row_idx, column=output_col).value
                if not existing_output or not str(existing_output).strip():
                    count += 1
                else:
                    out = str(existing_output)
                    # If previously marked as a blocked 400, skip it (do not retry)
                    if out.startswith("[ERROR]"):
                        if "400" in out or "Bad Request" in out or out.startswith("[BLOCKED]"):
                            # skip blocked entries
                            pass
                        else:
                            count += 1
    return count


def run_evaluation(input_path, model, provider, delay, output_path):
    # Initialize client based on provider
    if provider == "gemini":
        api_key = os.getenv("GEMINI_API_KEY")
        if not api_key:
            raise ValueError("GEMINI_API_KEY not set in environment or .env file")
        client_or_token = genai.Client(api_key=api_key)
    elif provider == "github":
        token = os.getenv("GITHUB_TOKEN")
        if not token:
            raise ValueError("GITHUB_TOKEN not set in environment or .env file")
        client_or_token = token
    else:
        raise ValueError(f"Unknown provider: {provider}")

    wb = openpyxl.load_workbook(input_path)

    grand_total = count_questions(wb)
    print(f"Total API calls to make: {grand_total}\n")

    call_num = 0
    total = 0
    errors = 0
    save_interval = 3  # Save file every 3 API calls

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            print(f"[skip] {sheet_name}")
            continue

        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        prompt_col = find_col(headers, PROMPT_HEADER)
        output_col = find_col(headers, OUTPUT_HEADER)

        if prompt_col is None:
            print(f"[warn] '{sheet_name}': column '{PROMPT_HEADER}' not found, skipping sheet")
            continue
        if output_col is None:
            print(f"[warn] '{sheet_name}': column '{OUTPUT_HEADER}' not found, skipping sheet")
            continue

        print(f"\n--- {sheet_name} ---")

        for row_idx in range(2, ws.max_row + 1):
            question = ws.cell(row=row_idx, column=prompt_col).value
            if not question or str(question).strip() == "":
                continue

            # Check if cell already has a successful response
            existing_output = ws.cell(row=row_idx, column=output_col).value
            if existing_output and str(existing_output).strip() and not str(existing_output).startswith("[ERROR]"):
                # Skip cells that already have successful responses
                continue

            call_num += 1
            question = str(question).strip()
            row_id = ws.cell(row=row_idx, column=1).value or f"row {row_idx}"
            preview = question[:72] + ("..." if len(question) > 72 else "")
            # Indicate if this is a retry
            retry_label = " (retry)" if existing_output else ""
            print(f"  [{call_num}/{grand_total}] {row_id}: {preview}{retry_label}")

            try:
                response = query_model(client_or_token, model, question, provider)
                ws.cell(row=row_idx, column=output_col).value = response
                print(f"    -> {len(response)} chars")
                total += 1
            except Exception as e:
                err_str = str(e)
                # Treat HTTP 400 / Bad Request as blocked by provider safety filter — skip retrying
                if "400" in err_str or "Bad Request" in err_str:
                    ws.cell(row=row_idx, column=output_col).value = f"[BLOCKED] {e}"
                    print(f"    -> BLOCKED: {e}")
                else:
                    ws.cell(row=row_idx, column=output_col).value = f"[ERROR] {e}"
                    print(f"    -> ERROR: {e}")
                errors += 1
                total += 1

            # Periodically save to prevent data loss on crash
            if call_num % save_interval == 0:
                wb.save(output_path)
                print(f"    [saved checkpoint]")

            if delay > 0:
                time.sleep(delay)

    wb.save(output_path)
    print(f"\n{'='*50}")
    print(f"Done. {total - errors}/{total} successful")
    print(f"Results saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Evaluate a model against the Teacher's Pet test suite."
    )
    parser.add_argument(
        "--input", required=True,
        help="Path to the evaluation xlsx file"
    )
    parser.add_argument(
        "--model", required=True,
        help="Model ID to test (e.g., gemini-2.5-flash-lite, gpt-4o, meta-llama-3.1-405b-instruct)"
    )
    parser.add_argument(
        "--provider", choices=["gemini", "github"], default="gemini",
        help="Model provider: gemini or github (default: gemini)"
    )
    parser.add_argument(
        "--delay", type=float, default=0.0,
        help="Seconds to wait between API calls (default: 0, auto-retries on rate limit)"
    )
    parser.add_argument(
        "--output", default=None,
        help="Output xlsx path (default: results_<provider>_<model>_<date>.xlsx)"
    )
    args = parser.parse_args()

    if args.output is None:
        today = date.today().isoformat()
        safe_model = args.model.replace("/", "-").replace(":", "-")
        args.output = f"results_{args.provider}_{safe_model}_{today}.xlsx"

    print(f"Input:    {args.input}")
    print(f"Provider: {args.provider}")
    print(f"Model:    {args.model}")
    print(f"Delay:    {args.delay}s between requests")
    print(f"Output:   {args.output}")
    print()

    run_evaluation(args.input, args.model, args.provider, args.delay, args.output)


if __name__ == "__main__":
    main()