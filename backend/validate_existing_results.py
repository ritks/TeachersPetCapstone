#!/usr/bin/env python3
"""
FIXED: Validation script for existing model evaluation results.

Reads an existing results xlsx file with responses already filled in,
validates each response using two GitHub models, and adds validation
columns to the spreadsheet.

This version properly detects rows where validators returned "error".

Usage:
    python validate_existing_results_v2.py --input validated_results_gemini-2.5-flash-lite_Engineered-Prompt.xlsx
"""

import argparse
import os
import time
from datetime import date

from dotenv import load_dotenv
import openpyxl
from rag.validator import ResponseValidator

load_dotenv()

# Sheet names to skip
SKIP_SHEETS = {"Instructions & Rubric", "Summary Dashboard"}

# Column headers
PROMPT_HEADER = "Test Input / Prompt"
OUTPUT_HEADER = "Actual Output"
VALIDATION_HEADER = "Validation Result"
VALIDATOR1_HEADER = "Validator 1"
VALIDATOR2_HEADER = "Validator 2"


def find_col(headers, name):
    """Return 1-based column index for a header name, or None if not found."""
    for i, h in enumerate(headers, 1):
        if h and str(h).strip() == name:
            return i
    return None


def needs_validation(ws, row_idx, validation_col, validator1_col, validator2_col):
    """Check if a row needs (re)validation by looking at validator columns."""
    if validation_col is None:
        return True
    
    existing_validation = ws.cell(row=row_idx, column=validation_col).value
    
    # No validation yet
    if not existing_validation or str(existing_validation).strip() == "":
        return True
    
    # Validation failed with error
    if str(existing_validation).startswith("ERROR"):
        return True
    
    # Check if either validator has "error" in it
    if validator1_col:
        v1 = ws.cell(row=row_idx, column=validator1_col).value
        if v1 and "error" in str(v1).lower():
            return True
    
    if validator2_col:
        v2 = ws.cell(row=row_idx, column=validator2_col).value
        if v2 and "error" in str(v2).lower():
            return True
    
    return False


def count_responses(wb):
    """Count total responses that need validation across all sheets."""
    count = 0
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        prompt_col = find_col(headers, PROMPT_HEADER)
        output_col = find_col(headers, OUTPUT_HEADER)
        validation_col = find_col(headers, VALIDATION_HEADER)
        validator1_col = find_col(headers, VALIDATOR1_HEADER)
        validator2_col = find_col(headers, VALIDATOR2_HEADER)
        
        if prompt_col is None or output_col is None:
            continue
        
        for row_idx in range(2, ws.max_row + 1):
            prompt = ws.cell(row=row_idx, column=prompt_col).value
            response = ws.cell(row=row_idx, column=output_col).value
            
            # Count if there's a response and validation is needed
            if response and str(response).strip():
                if needs_validation(ws, row_idx, validation_col, validator1_col, validator2_col):
                    count += 1
    return count


def validate_results(input_path, output_path, delay=1.0):
    """Validate existing responses in the Excel file."""
    
    # Initialize validator
    try:
        validator = ResponseValidator()
        print("[INFO] Validator initialized\n")
    except ValueError as e:
        raise ValueError(f"Failed to initialize validator: {e}")
    
    # Load workbook
    wb = openpyxl.load_workbook(input_path)
    
    total_responses = count_responses(wb)
    print(f"Total responses to validate: {total_responses}\n")
    
    if total_responses == 0:
        print("All responses already validated successfully!")
        return
    
    validated = 0
    errors = 0
    save_interval = 5  # Save every 5 validations
    
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            print(f"[skip] {sheet_name}")
            continue
        
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        
        prompt_col = find_col(headers, PROMPT_HEADER)
        output_col = find_col(headers, OUTPUT_HEADER)
        
        if prompt_col is None or output_col is None:
            print(f"[warn] '{sheet_name}': required columns not found, skipping")
            continue
        
        # Find or create validation columns
        validation_col = find_col(headers, VALIDATION_HEADER)
        validator1_col = find_col(headers, VALIDATOR1_HEADER)
        validator2_col = find_col(headers, VALIDATOR2_HEADER)
        
        if validation_col is None:
            # Add validation columns at the end
            last_col = len([h for h in headers if h]) + 1
            validation_col = last_col
            validator1_col = last_col + 1
            validator2_col = last_col + 2
            
            ws.cell(row=1, column=validation_col).value = VALIDATION_HEADER
            ws.cell(row=1, column=validator1_col).value = VALIDATOR1_HEADER
            ws.cell(row=1, column=validator2_col).value = VALIDATOR2_HEADER
        
        print(f"\n--- {sheet_name} ---")
        
        for row_idx in range(2, ws.max_row + 1):
            prompt = ws.cell(row=row_idx, column=prompt_col).value
            response = ws.cell(row=row_idx, column=output_col).value
            
            if not response or not str(response).strip():
                continue
            
            # Skip if already validated successfully (checks validator columns for errors)
            if not needs_validation(ws, row_idx, validation_col, validator1_col, validator2_col):
                continue
            
            question = str(prompt).strip()
            answer = str(response).strip()
            
            row_id = ws.cell(row=row_idx, column=1).value or f"row {row_idx}"
            preview = question[:60] + ("..." if len(question) > 60 else "")
            
            # Check if this is a retry
            existing_validation = ws.cell(row=row_idx, column=validation_col).value
            retry_label = " (retry)" if existing_validation else ""
            
            print(f"  [{validated + errors + 1}/{total_responses}] {row_id}: {preview}{retry_label}")
            
            try:
                validation_result = validator.validate(question, answer)
                is_safe = validation_result["is_safe"]
                votes = validation_result["safety_votes"]
                
                # Extract individual validator votes
                validator_names = list(votes.keys())
                vote1 = votes.get(validator_names[0], "error") if len(validator_names) > 0 else "N/A"
                vote2 = votes.get(validator_names[1], "error") if len(validator_names) > 1 else "N/A"
                
                # Write validation results
                ws.cell(row=row_idx, column=validation_col).value = "SAFE" if is_safe else "UNSAFE"
                ws.cell(row=row_idx, column=validator1_col).value = f"{validator_names[0]}: {vote1}"
                ws.cell(row=row_idx, column=validator2_col).value = f"{validator_names[1]}: {vote2}"
                
                print(f"    -> {'SAFE' if is_safe else 'UNSAFE'} [{vote1}, {vote2}]")
                validated += 1
                
            except Exception as e:
                print(f"    -> ERROR: {e}")
                ws.cell(row=row_idx, column=validation_col).value = f"ERROR: {str(e)[:100]}"
                errors += 1
            
            # Periodic save
            if (validated + errors) % save_interval == 0:
                wb.save(output_path)
                print(f"    [saved checkpoint]")
            
            # Rate limiting delay
            if delay > 0:
                time.sleep(delay)
    
    # Final save
    wb.save(output_path)
    print(f"\n{'='*60}")
    print(f"Validation complete!")
    print(f"  Successfully validated: {validated}/{total_responses}")
    print(f"  Errors: {errors}")
    print(f"Results saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Validate existing model evaluation results."
    )
    parser.add_argument(
        "--input", required=True,
        help="Path to existing results xlsx file"
    )
    parser.add_argument(
        "--output", default=None,
        help="Output xlsx path (default: same as input)"
    )
    parser.add_argument(
        "--delay", type=float, default=5.0,
        help="Seconds to wait between validation calls (default: 5.0)"
    )
    args = parser.parse_args()
    
    if args.output is None:
        args.output = args.input  # Overwrite input file
    
    print(f"Input:  {args.input}")
    print(f"Output: {args.output}")
    print(f"Delay:  {args.delay}s between validations")
    print()
    
    validate_results(args.input, args.output, args.delay)


if __name__ == "__main__":
    main()
